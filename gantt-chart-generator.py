import pandas as pd
from openpyxl.styles import PatternFill
import datetime
import re
import shutil
import calendar


class Timeline:
    def __init__(self):
        self.df = self.load_file()
        self.file = self.get_filename()


    def get_filename(self):
        '''
        Output file name
        '''
        self.file = f"Monthly Timeline {calendar.month_name[int(self.month)]} {self.year}.xlsx"


    def copy_excel(self):
        '''
        Create a copy of the 'monthly-schedule.xlsx' input file
        '''
        shutil.copy('monthly-schedule.xlsx', self.file)


    def load_file(self):
        '''
        Load excel file into df
        '''
        return pd.read_excel('monthly-schedule.xlsx', dtype=object, usecols='B:D')


    def get_period(self):
        '''
        Get year and month from user input
        '''
        self.period = input("Input Year-Month (ex. 09-2024) : ")
        match = re.match(r"(\d{2})-(\d{4})", self.period.strip())
        if match:
            self.month = match.group(1)
            self.year = match.group(2)


    def add_dates_col(self, df):
        '''
        Add columns for each day of the month (1-31) to df
        Creates list containing dates of the month that are weekends

        Takes in df, adds 31 columns (dates), returns df
        '''
        self.weekend_days = []
        for day in range(1, 32):
            df[str(day)] = ''
            try:
                thisdate = datetime.date(int(self.year), int(self.month), day)
            except(ValueError):
                break
            if thisdate.weekday() > 4:
                self.weekend_days.append(day)
        return df
        

    def construct_df(self):
        '''
        Calls add_dates_col and creates variable 'df_temp'
        '''
        self.df_temp = self.add_dates_col(self.df)
        

    def add_start_end_date(self):
        '''
        Set the 'start_date' and 'end_date' columns in df to datetime format
        Note: all rows for start_date and end_date need to be filled 
        ''' 
        self.df_temp['Start Date'] = pd.to_datetime(self.df_temp['Start Date'].astype(str)+ f'-{self.month}-{self.year}', format='%d-%m-%Y').dt.strftime("%A, %d %B %Y")
        self.df_temp['End Date'] = pd.to_datetime(self.df_temp['End Date'].astype(str)+ f'-{self.month}-{self.year}', format='%d-%m-%Y').dt.strftime("%A, %d %B %Y")


    def mark_tasks(self):
        '''
        Mark the date columns (1-31) for each task (row) according to the start and end dates
        Creates a gantt chart for all tasks
        '''
        for idx, row in self.df_temp.iterrows():
            start_date = pd.to_datetime(row['start_date']).day
            end_date = pd.to_datetime(row['end_date']).day
            for day in range(start_date, end_date + 1):
                self.df_temp.loc[idx, str(day)] = ' '  # Task marker
    

    def color_task(self):
        '''
        Apply color fill (green) to the cells with ' ' (gantt chart)
        Green color will fill the gantt chart according to tasks' start and end date
        '''
        fill_task = PatternFill(start_color="BEE7A5", end_color="BEE7A5", fill_type="solid")
        for row in self.worksheet.iter_rows(min_row=2, min_col=5, max_col=35):  # Date columns start at column index 5
            for cell in row:
                if cell.value == ' ':
                    cell.fill = fill_task
    

    def color_date(self):
        '''
        Apply color fill (red) to dates (row 1) that are weekends
        Red color will fill the dates that are weekends (row 1)
        '''
        fill_weekend = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for row in self.worksheet.iter_rows(min_row=1, max_row=1, min_col=5, max_col=35):  # Date row on row 1
            for cell in row:
                if int(cell.value) in self.weekend_days:
                    cell.fill = fill_weekend


    def adjust_column_width(self):
        '''
        Adjusts the column width
        '''
        for column in self.worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 3)
                self.worksheet.column_dimensions[column_letter].width = adjusted_width


    def create_excel(self):
        # Create copy of input schedule document (excel file)
        self.copy_excel()

        # Save the DataFrame to Excel (copied document)
        file_name = self.file
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            self.df_temp.to_excel(writer , index=False, sheet_name="Sheet1", startcol=1)

            # Access the worksheet to apply formatting
            workbook = writer.book
            self.worksheet = workbook["Sheet1"]

            # Apply color fill (green) to the cells with ' ' (gantt chart)
            self.color_task()

            # Apply color fill (red) to dates (row 1) that are weekends
            self.color_date()
            
            # Adjust column width
            self.adjust_column_width()
    

    def end_to_end(self):
        self.get_period()
        self.get_filename()
        self.construct_df()
        self.add_start_end_date()
        self.mark_tasks()
        self.create_excel()


def main():
    timeline = Timeline()
    timeline.end_to_end()


if __name__ == "__main__":
    main()
